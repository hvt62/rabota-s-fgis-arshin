import os
import time
import threading
import uuid
import requests
import openpyxl
from io import BytesIO
from flask import Flask, render_template, request, send_file, redirect, url_for

app = Flask(__name__)

VERSION = "1.0.0"

# Новый рабочий endpoint ФГИС «Аршин»
ARSHIN_URL = "https://fgis.gost.ru/fundmetrology/cm/xcdb/vri/select"
ARSHIN_MAIN = "https://fgis.gost.ru/fundmetrology/cm/"

# Поля, которые запрашиваем у API
FIELDS = [
    "vri_id",
    "org_title",
    "mi.mitnumber",
    "mi.mititle",
    "mi.mitype",
    "mi.modification",
    "mi.number",
    "verification_date",
    "valid_date",
    "applicability",
    "result_docnum",
    "sticker_num",
]

# Сессия для запросов к Аршину (с cookies)
_session = None
_session_lock = threading.Lock()

# Хранилище прогресса задач
progress_store = {}
progress_lock = threading.Lock()


def get_arshin_session():
    """Создаёт сессию с cookies, полученными с главной страницы Аршина."""
    global _session
    if _session is not None:
        return _session

    with _session_lock:
        if _session is not None:
            return _session

        session = requests.Session()
        session.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "ru,en;q=0.9",
        })

        try:
            resp = session.get(ARSHIN_MAIN, timeout=30)
            resp.raise_for_status()
            print(f"[Сессия] Статус: {resp.status_code}")
            print(f"[Сессия] Cookies от сервера: {dict(session.cookies)}")
        except Exception as e:
            print(f"[Сессия] Ошибка получения cookies: {e}")

        _session = session
        return _session


def _do_request(mi_number, params, label):
    """Выполнить запрос к API с 3 попытками при 500."""
    headers = {"Referer": "https://fgis.gost.ru/fundmetrology/cm/results"}
    session = get_arshin_session()

    for attempt in range(1, 4):
        try:
            resp = session.get(ARSHIN_URL, params=params, headers=headers, timeout=60)
            print(f"[API] {mi_number}: {label}, попытка {attempt}, статус {resp.status_code}")
            resp.raise_for_status()
            data = resp.json()
            docs = data.get("response", {}).get("docs", [])
            print(f"[API] {mi_number}: {label}: найдено {len(docs)} документов")
            return docs
        except requests.exceptions.HTTPError as e:
            status = e.response.status_code if e.response is not None else 0
            if status == 500 and attempt < 3:
                print(f"[API] {mi_number}: {label}: 500, пауза 7.5с, попытка {attempt+1}/3")
                time.sleep(7.5)
                continue
            print(f"[API] {mi_number}: {label}: HTTP ошибка {status}")
            return {"error": str(e)}
        except requests.exceptions.RequestException as e:
            print(f"[API] {mi_number}: {label}: ошибка {e}")
            if attempt < 3:
                time.sleep(7.5)
                continue
            return {"error": str(e)}

    return {"error": "3 попытки не удались"}


def search_arshin(mi_number, mi_type=None):
    """Поиск по подстроке mi.number:*номер* + mi.modification:*тип* (если задан).
       3 попытки при 500 ошибке."""
    fq = [f"mi.number:*{mi_number}*"]
    if mi_type:
        fq.append(f"mi.modification:*{mi_type}*")
    params = {
        "fq": fq,
        "q": "*",
        "fl": ",".join(FIELDS),
        "sort": "verification_date desc,org_title asc",
        "rows": 1000,
        "start": 0,
    }
    return _do_request(mi_number, params, "поиск")


def format_date(date_str):
    if not date_str:
        return ""
    return date_str[:10].replace("-", ".")


def format_applicability(val):
    if val is True:
        return "ГОДЕН"
    elif val is False:
        return "НЕ ГОДЕН"
    return str(val)


def process_docs(docs, num, mi_type):
    """Обработать список документов: форматирование без локальной фильтрации."""
    if not docs:
        return []

    records = []
    for doc in docs:
        record = {
            "number": num,
            "input_type": mi_type,
            "mi_number": doc.get("mi.number", ""),
            "title": doc.get("mi.mititle", ""),
            "type": doc.get("mi.mitype", ""),
            "modification": doc.get("mi.modification", ""),
            "verification_date": format_date(doc.get("verification_date", "")),
            "valid_date": format_date(doc.get("valid_date", "")),
            "applicability": format_applicability(doc.get("applicability")),
            "org_title": doc.get("org_title", ""),
            "result_docnum": doc.get("result_docnum", ""),
        }
        records.append(record)

    return records


def cancellable_sleep(task_id, seconds):
    """Сон с проверкой отмены каждую секунду."""
    for _ in range(int(seconds)):
        with progress_lock:
            if progress_store[task_id].get("cancel"):
                return True
        time.sleep(1)
    return False


def process_items_background(task_id, rows_data):
    """Фоновая обработка всех номеров с обновлением прогресса."""
    delays = [3, 7, 12, 18, 25]
    pending = list(rows_data)
    errors_list = []

    with progress_lock:
        progress_store[task_id]["pending_items"] = pending

    for iteration, delay in enumerate(delays, 1):
        if not pending:
            break

        with progress_lock:
            progress_store[task_id]["errors"] = 0
            progress_store[task_id]["not_found"] = 0

        print(f"\n[App] === Итерация {iteration}, пауза {delay}с, осталось {len(pending)} номеров ===")

        still_pending = []

        for idx, item in enumerate(pending):
            with progress_lock:
                if progress_store[task_id].get("cancel"):
                    print(f"[App] {task_id}: получен сигнал отмены, завершаем поток")
                    return

            num = item["number"]
            mi_type = item["type"] if item["type"] else ""

            docs = search_arshin(num, mi_type if mi_type else None)

            if isinstance(docs, dict) and "error" in docs:
                still_pending.append(item)
                with progress_lock:
                    progress_store[task_id]["errors"] += 1
                if idx < len(pending) - 1:
                    if cancellable_sleep(task_id, delay):
                        return
                continue

            if not docs:
                still_pending.append(item)
                with progress_lock:
                    progress_store[task_id]["not_found"] += 1
                if idx < len(pending) - 1:
                    if cancellable_sleep(task_id, delay):
                        return
                continue

            records = process_docs(docs, num, mi_type)

            with progress_lock:
                progress_store[task_id]["results"].extend(records)
                progress_store[task_id]["processed"] += 1
                pending_items = progress_store[task_id].get("pending_items", [])
                progress_store[task_id]["pending_items"] = [pi for pi in pending_items if pi["number"] != num]

            if idx < len(pending) - 1:
                if cancellable_sleep(task_id, delay):
                    return

        pending = still_pending

        with progress_lock:
            progress_store[task_id]["pending_items"] = pending

    with progress_lock:
        pending = progress_store[task_id].get("pending_items", [])
        results = progress_store[task_id]["results"]

    for item in pending:
        results.append({
            "number": item["number"],
            "input_type": item["type"] if item["type"] else "",
            "mi_number": "",
            "title": "Не найдено",
            "type": "",
            "modification": "",
            "verification_date": "",
            "valid_date": "",
            "applicability": "",
            "org_title": "",
            "result_docnum": "",
        })

    order = {item["number"]: idx for idx, item in enumerate(rows_data)}
    results.sort(key=lambda r: order.get(r["number"], 999))

    with progress_lock:
        progress_store[task_id]["status"] = "complete"
        progress_store[task_id]["results"] = results
        progress_store[task_id]["errors_list"] = errors_list
        progress_store[task_id]["not_found_count"] = sum(1 for r in results if r["title"] == "Не найдено")
        progress_store[task_id]["rows_data"] = rows_data


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        mode = request.form.get("mode", "file")

        # --- Ручной ввод ---
        if mode == "manual":
            manual_number = request.form.get("manual_number", "").strip()
            if not manual_number:
                return render_template("index.html", error="Введите номер СИ")

            manual_type = request.form.get("manual_type", "").strip()
            rows_data = [{"number": manual_number, "type": manual_type}]

            print(f"[App] Ручной ввод: номер={manual_number}, тип={manual_type}")

        # --- Загрузка файла ---
        else:
            file = request.files.get("file")
            if not file or file.filename == "":
                return render_template("index.html", error="Файл не выбран")

            try:
                wb = openpyxl.load_workbook(file, read_only=True)
                ws = wb.active
            except Exception as e:
                return render_template("index.html", error=f"Ошибка чтения Excel: {e}")

            rows_data = []
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                num = row[0]
                if num is not None:
                    num = str(num).strip()
                    type_val = str(row[1]).strip() if row[1] is not None else ""
                    rows_data.append({"number": num, "type": type_val})

            print(f"[App] Прочитано строк: {len(rows_data)}")

            if not rows_data:
                return render_template("index.html", error="Нет номеров в первом столбце")

        # --- Общий запуск задачи ---
        task_id = str(uuid.uuid4())
        with progress_lock:
            progress_store[task_id] = {
                "total": len(rows_data),
                "processed": 0,
                "errors": 0,
                "not_found": 0,
                "status": "running",
                "cancel": False,
                "start_time": time.time(),
                "results": [],
                "errors_list": [],
                "rows_data": rows_data,
                "not_found_count": 0,
                "pending_items": [],
            }

        thread = threading.Thread(target=process_items_background, args=(task_id, rows_data))
        thread.daemon = True
        thread.start()

        # Ручной ввод — сразу на страницу результатов (с ожиданием)
        if mode == "manual":
            return redirect(url_for("results_page", task_id=task_id))

        return redirect(url_for("progress_page", task_id=task_id))

    return render_template("index.html")


@app.route("/progress/<task_id>")
def progress_page(task_id):
    with progress_lock:
        if task_id not in progress_store:
            return redirect(url_for("index"))
    return render_template("progress.html", task_id=task_id, version=VERSION)


@app.route("/progress_data/<task_id>")
def progress_data(task_id):
    with progress_lock:
        data = progress_store.get(task_id)
        if not data:
            return {"status": "error", "message": "Задача не найдена"}
        return {
            "total": data["total"],
            "processed": data["processed"],
            "errors": data["errors"],
            "not_found": data["not_found"],
            "status": data["status"],
        }


@app.route("/cancel_task/<task_id>", methods=["POST"])
def cancel_task(task_id):
    with progress_lock:
        data = progress_store.get(task_id)
        if not data or data["status"] != "running":
            return {"status": "error", "message": "Задача не найдена или уже завершена"}
        data["cancel"] = True

        rows_data = data["rows_data"]
        results = data["results"]
        errors_list = data["errors_list"]
        pending = data.get("pending_items", [])

        for item in pending:
            results.append({
                "number": item["number"],
                "input_type": item["type"] if item["type"] else "",
                "mi_number": "",
                "title": "Не найдено",
                "type": "",
                "modification": "",
                "verification_date": "",
                "valid_date": "",
                "applicability": "",
                "org_title": "",
                "result_docnum": "",
            })

        order = {item["number"]: idx for idx, item in enumerate(rows_data)}
        results.sort(key=lambda r: order.get(r["number"], 999))

        data["status"] = "complete"
        data["results"] = results
        data["errors_list"] = errors_list
        data["not_found_count"] = sum(1 for r in results if r["title"] == "Не найдено")
        data["rows_data"] = rows_data

    return {"status": "ok"}


@app.route("/results/<task_id>")
def results_page(task_id):
    with progress_lock:
        data = progress_store.get(task_id)
        if not data:
            return redirect(url_for("index"))

        # Если задача ещё выполняется — показываем страницу с ожиданием
        if data["status"] != "complete":
            return render_template("result.html", task_id=task_id, waiting=True,
                                   total=0, total_records=0, not_found_count=0,
                                   results=[], errors=[], page=1, total_pages=1,
                                   per_page=50, all_results=[])

        results = data["results"]
        rows_data = data["rows_data"]
        errors = data["errors_list"]
        not_found_count = data["not_found_count"]
        total_records = len(results) - not_found_count

    per_page = 50
    total_pages = max(1, (len(results) + per_page - 1) // per_page)
    page_results = results[:per_page]

    return render_template(
        "result.html", results=page_results, errors=errors, total=len(rows_data),
        not_found_count=not_found_count,
        page=1, total_pages=total_pages, per_page=per_page,
        all_results=results, total_records=total_records,
        task_id=task_id, waiting=False
    )


@app.route("/page/<task_id>", methods=["POST"])
def page(task_id):
    with progress_lock:
        data = progress_store.get(task_id)
        if not data or data["status"] != "complete":
            return redirect(url_for("index"))

        results = data["results"]

    page = int(request.form.get("page", 1))
    per_page = 50
    total_pages = max(1, (len(results) + per_page - 1) // per_page)

    if page < 1:
        page = 1
    if page > total_pages:
        page = total_pages

    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    page_results = results[start_idx:end_idx]

    return render_template(
        "result.html", results=page_results, errors=[], total=0,
        not_found_count=0,
        page=page, total_pages=total_pages, per_page=per_page,
        all_results=results, total_records=len(results),
        task_id=task_id, waiting=False
    )


@app.route("/download/<task_id>", methods=["POST"])
def download(task_id):
    with progress_lock:
        data = progress_store.get(task_id)
        if not data or data["status"] != "complete":
            return {"status": "error", "message": "Результаты не найдены"}

        results = data["results"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты"

    headers = [
        "Искомый номер",
        "Заданный тип",
        "Заводской номер СИ",
        "Наименование СИ",
        "Тип СИ",
        "Модификация",
        "Дата поверки",
        "Действителен до",
        "Результат",
        "Организация",
        "Номер документа",
    ]
    ws.append(headers)

    for r in results:
        ws.append([
            r.get("number", ""),
            r.get("input_type", ""),
            r.get("mi_number", ""),
            r.get("title", ""),
            r.get("type", ""),
            r.get("modification", ""),
            r.get("verification_date", ""),
            r.get("valid_date", ""),
            r.get("applicability", ""),
            r.get("org_title", ""),
            r.get("result_docnum", ""),
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="results.xlsx",
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)